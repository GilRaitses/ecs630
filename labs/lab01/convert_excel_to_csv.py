#!/usr/bin/env python3
"""
Convert Arena Excel workbook (.xlsm) to CSV files (one per sheet)
"""
import sys
import os

def convert_excel_to_csvs(excel_file):
    """Convert each sheet in Excel file to separate CSV"""
    
    # Try using openpyxl first
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheets = wb.sheetnames
        
        print(f"Found {len(sheets)} sheets:")
        for i, sheet_name in enumerate(sheets, 1):
            print(f"  {i}. {sheet_name}")
        
        # Output directory
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_dir = f"{base_name}_csvs"
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
            
            with open(csv_file, 'w', encoding='utf-8') as f:
                for row in sheet.iter_rows(values_only=True):
                    # Skip empty rows completely
                    if any(cell is not None and str(cell).strip() for cell in row):
                        # Write row with proper CSV formatting
                        csv_row = ','.join([
                            '"' + str(cell).replace('"', '""') + '"' 
                            if cell is not None else ''
                            for cell in row
                        ])
                        f.write(csv_row + '\n')
            
            print(f"  ✓ Created: {csv_file}")
        
        print(f"\n✓ Conversion complete! Files saved in: {output_dir}/")
        return True
        
    except ImportError:
        print("openpyxl not available, trying alternative methods...")
        
    # Try using xlrd (legacy Excel format)
    try:
        import xlrd
        wb = xlrd.open_workbook(excel_file)
        sheets = wb.sheet_names
        
        print(f"Found {len(sheets)} sheets:")
        for i, sheet_name in enumerate(sheets, 1):
            print(f"  {i}. {sheet_name}")
        
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_dir = f"{base_name}_csvs"
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet_name in sheets:
            sheet = wb.sheet_by_name(sheet_name)
            csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
            
            with open(csv_file, 'w', encoding='utf-8') as f:
                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    if any(cell for cell in row):
                        csv_row = ','.join([
                            '"' + str(cell).replace('"', '""') + '"' 
                            for cell in row
                        ])
                        f.write(csv_row + '\n')
            
            print(f"  ✓ Created: {csv_file}")
        
        print(f"\n✓ Conversion complete! Files saved in: {output_dir}/")
        return True
        
    except ImportError:
        print("xlrd not available.")
    
    # Last resort: use csvkit if available
    try:
        print("\nTrying csvkit...")
        import subprocess
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_dir = f"{base_name}_csvs"
        os.makedirs(output_dir, exist_ok=True)
        
        result = subprocess.run(
            ['in2csv', excel_file, '--names'],
            capture_output=True,
            text=True
        )
        
        if result.returncode == 0:
            sheet_names = result.stdout.strip().split('\n')
            print(f"Found {len(sheet_names)} sheets")
            
            for sheet_name in sheet_names:
                csv_file = os.path.join(output_dir, f"{sheet_name}.csv")
                subprocess.run(
                    ['in2csv', excel_file, '--sheet', sheet_name],
                    stdout=open(csv_file, 'w')
                )
                print(f"  ✓ Created: {csv_file}")
            
            print(f"\n✓ Conversion complete! Files saved in: {output_dir}/")
            return True
    except:
        pass
    
    print("\n❌ Error: No suitable Excel reader found.")
    print("Install one of: openpyxl, xlrd, or csvkit")
    print("  pip install openpyxl")
    return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert_excel_to_csv.py <excel_file.xlsm>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)
    
    convert_excel_to_csvs(excel_file)

